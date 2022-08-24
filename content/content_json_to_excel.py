#!/usr/bin/env python

"""
"""
import csv
import json
from pathlib import Path
import string
import sys

import openpyxl

from census_objects import CensusTopic, topic_from_content_json


def load_content(content_file: Path) -> list[CensusTopic]:
    """Load all census objects defined in content_file."""
    with open(content_file, "r") as f:
        raw_topics = json.load(f)
    return [topic_from_content_json(t) for t in raw_topics]


def iterate_column_ref(column_ref: str) -> str:
    if column_ref.endswith("Z"):
        return "A" * (len(column_ref) + 1)
    letters = string.ascii_letters
    next_letter = letters[letters.index(column_ref[-1]) + 1]
    return column_ref[:-1] + next_letter


def main(content_json_fn: Path, output_dir: Path, index_csv_fn: Path or None) -> None:
    all_topics = load_content(content_json_fn)
    wb = openpyxl.Workbook()

    # add index page
    index_ws = wb.active
    index_ws.title = "Index"
    index_ws.cell(row=1, column=1, value="Release")
    index_ws.cell(row=1, column=2, value="Topic Grouping")
    index_ws.cell(row=1, column=3, value="Topic")
    index_ws.cell(row=1, column=4, value="Variable")
    index_ws.cell(row=1, column=5, value="Classifications To Include (leave blank if none)")
    index_ws.cell(row=1, column=6, value="Choropleth Default Classification (required)")
    index_ws.cell(row=1, column=7, value="Dot Density Default Classification (leave blank if none)")
    index_row = 2

    # load info from pre-baked index csv is one if provided
    if index_csv_fn is not None:
        with open(index_csv_fn, "r") as f:
            index_from_csv = list(csv.DictReader(f))

    # add variable pages with all classifications
    for topic in all_topics:
        for variable in topic.variables:
            ws = wb.create_sheet(variable.code)

            # link to index page
            index_ws.cell(row=index_row, column=3, value=topic.name)
            index_ws.cell(row=index_row, column=4).hyperlink=f"#{variable.code}!A1"
            index_ws.cell(row=index_row, column=4).value=variable.code
            index_ws.cell(row=index_row, column=4).style="Hyperlink"

            # append data from pre-baked index rows if the variable is referenced there
            if index_csv_fn is not None:
                index_from_csv_row = next((r for r in index_from_csv if r["Variable"] == variable.code), None)
                if index_from_csv_row is not None:
                    index_ws.cell(row=index_row, column=5, value=index_from_csv_row["Classifications To Include (leave blank if none)"])
                    index_ws.cell(row=index_row, column=6, value=index_from_csv_row["Choropleth Default Classification (required)"])
                    index_ws.cell(row=index_row, column=7, value=index_from_csv_row["Dot Density Default Classification (leave blank if none)"])

            index_row += 1
            column = 1
            row = 1

            for classification in variable.classifications:
                ws.cell(row=row, column=column, value=classification.code)
                row +=1
                ws.cell(row=row, column=column, value="")
                row +=1
                ws.cell(row=row, column=column, value="Categories:")
                row +=1
                for category in classification.categories:
                    ws.cell(row=row, column=column, value=category.name)
                    row +=1
                column +=2
                row = 1

    output_filename = output_dir.joinpath(f"{content_json_fn.stem}.xlsx")
    wb.save(output_filename)



if __name__ == "__main__":
    content_json_fn = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    if len(sys.argv) == 4:
       index_csv_fn = Path(sys.argv[3])
    else:
        index_csv_fn = None
    main(content_json_fn, output_dir, index_csv_fn)
